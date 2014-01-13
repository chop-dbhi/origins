from __future__ import division, unicode_literals, absolute_import

from . import base, _file

import openpyxl


class Client(_file.Client):
    def __init__(self, path, headers=True, **kwargs):
        super(Client, self).__init__(path, **kwargs)

        self._sheet_columns = {}

        if isinstance(headers, bool):
            self.has_headers = headers
        else:
            self.has_headers = False
            first_sheet = self.workbook.get_sheet_names()[0]
            if isinstance(headers, (list, tuple)):
                self._sheet_columns[first_sheet] = headers
            elif isinstance(headers, dict):
                self._sheet_columns = headers
            else:
                raise TypeError('headers must be a bool, list or dict')

    @property
    def workbook(self):
        return openpyxl.load_workbook(self.file_path, use_iterators=True)

    def sheets(self):
        return [{
            'name': name,
            'index': i,
        } for i, name in enumerate(self.workbook.get_sheet_names())]

    def columns(self, sheet_name):
        "Return the sheet header names or indices."
        if sheet_name in self._sheet_columns:
            column_names = self._sheet_columns[sheet_name]
        else:
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            first_row = next(sheet.iter_rows())
            if self.has_headers:
                column_names = [c.internal_value for c in first_row]
            else:
                column_names = range(len(first_row))

        columns = []

        for i, name in enumerate(column_names):
            columns.append({
                'name': name,
                'index': i
            })

        return columns


class Workbook(_file.File):
    def sync(self):
        super(Workbook, self).sync()
        self.define(self.client.sheets(), Sheet)

    @property
    def sheets(self):
        return self.definitions('sheet', sort='index')


class Sheet(base.Node):
    def sync(self):
        self.define(self.client.columns(self['name']), Column)

    @property
    def columns(self):
        return self.definitions('column', sort='index')


class Column(base.Node):
    pass


# Export for API
Origin = Workbook
